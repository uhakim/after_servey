from io import BytesIO
import datetime as dt
import os
import re

import openpyxl
from openpyxl.cell.cell import MergedCell


def _build_default_roster_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "4-4"

    ws.merge_cells("A1:J1")
    ws["A1"] = "2026학년도 동성초등학교 학생일람표"
    ws.merge_cells("A2:B2")
    ws.merge_cells("C2:G2")
    ws["A2"] = "남"
    ws["C2"] = "    4학년      4반     담임  : "

    ws["A3"] = "번\n호"
    ws["B3"] = "성   명"
    ws["C3"] = "생년월일"
    ws["D3"] = "주소(도로명주소)"
    ws["E3"] = "보호자"
    ws["G3"] = "연락처"
    ws["I3"] = "형제\n관계"
    ws["J3"] = "학교\n등교차"
    ws["E4"] = "부"
    ws["F4"] = "모"
    ws["G4"] = "부"
    ws["H4"] = "모"

    ws.merge_cells("A25:B25")
    ws["A25"] = "여"

    for n in range(1, 21):
        ws.cell(4 + n, 1).value = n
    for n in range(41, 61):
        ws.cell(n - 13, 1).value = n

    for c, w in [(1, 6), (2, 10), (3, 14), (4, 42), (5, 10), (6, 10), (7, 14), (8, 14), (9, 14), (10, 12)]:
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    return wb


def _load_workbook(template_bytes=None, default_path=None):
    if template_bytes:
        return openpyxl.load_workbook(BytesIO(template_bytes))
    if default_path and os.path.exists(default_path):
        return openpyxl.load_workbook(default_path)
    return _build_default_roster_template()


def _replace_school_year(ws, school_year):
    pattern = re.compile(r"\d{4}학년도")
    for r in range(1, 8):
        for c in range(1, 12):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "학년도" in v:
                ws.cell(r, c).value = pattern.sub(f"{school_year}학년도", v)


def _parse_int(v):
    if v in (None, ""):
        return None
    m = re.search(r"\d+", str(v))
    if not m:
        return None
    try:
        return int(m.group(0))
    except Exception:
        return None


def _normalize_address(addr):
    s = str(addr or "").strip()
    if not s:
        return ""
    s = re.sub(r"\s+", " ", s)
    s = s.replace("부산시 ", "부산광역시 ")
    s = s.replace("부산시", "부산광역시")
    if re.match(r"^[가-힣]+구\b", s):
        s = "부산광역시 " + s
    return s


def _normalize_birth_with_fallback(raw, parsed):
    raw_s = str(raw or "").strip()
    if isinstance(parsed, dt.date):
        return parsed

    digits = re.sub(r"\D", "", raw_s)
    candidate = None
    try:
        if len(digits) == 8:
            y, m, d = int(digits[:4]), int(digits[4:6]), int(digits[6:8])
            candidate = dt.date(y, m, d)
        elif len(digits) == 6:
            y, m, d = int(digits[:2]) + 2000, int(digits[2:4]), int(digits[4:6])
            candidate = dt.date(y, m, d)
    except Exception:
        candidate = None
    return candidate


def _format_birth_iso(d):
    return f"{d.year:04d}-{d.month:02d}-{d.day:02d}"


def _clean_siblings(value):
    s = str(value or "").strip()
    if not s:
        return ""
    compact = re.sub(r"[\s.,]+", "", s)
    if compact in {"없음", "해당없음", "없다", "없습니다", "무"}:
        return ""
    if re.fullmatch(r"(없음)+", compact) or re.fullmatch(r"(없습니다)+", compact):
        return ""
    return s


def _assign_by_number(records, allowed_slots):
    mapped = {}
    used = set()
    slot_set = set(allowed_slots)

    for rec in sorted(records, key=lambda x: (x.get("number") or 9999, x.get("name", ""))):
        n = rec.get("number")
        if n in slot_set and n not in used:
            mapped[n] = rec
            used.add(n)

    remaining_records = [r for r in records if r not in mapped.values()]
    remaining_slots = [s for s in allowed_slots if s not in used]

    for rec, slot in zip(sorted(remaining_records, key=lambda x: (x.get("name", ""))), remaining_slots):
        mapped[slot] = rec

    return mapped


def _safe_set(ws, row, col, value):
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def build_student_roster(records, school_year, template_bytes=None, default_template_path=None):
    wb = _load_workbook(template_bytes, default_template_path)
    ws = wb["4-4"] if "4-4" in wb.sheetnames else wb.active

    _replace_school_year(ws, school_year)

    if int(school_year) % 2 == 0:
        ws.cell(2, 1).value = "남"
        ws.cell(25, 1).value = "여"
    else:
        ws.cell(2, 1).value = "여"
        ws.cell(25, 1).value = "남"

    if records:
        grade = records[0].get("grade_num") or 0
        class_ = records[0].get("class_num") or 0
        ws.cell(2, 3).value = f"    {grade}학년      {class_}반     담임  : "
        ws.cell(2, 8).value = None

    num_to_row = {}
    for r in range(1, ws.max_row + 1):
        n = _parse_int(ws.cell(r, 1).value)
        if n is not None:
            num_to_row[n] = r
            for c in range(2, 11):
                _safe_set(ws, r, c, None)

    all_slots = sorted(num_to_row.keys())
    mapped = _assign_by_number(records, all_slots)

    for slot, rec in mapped.items():
        r = num_to_row[slot]
        _safe_set(ws, r, 1, slot)
        _safe_set(ws, r, 2, rec.get("name", ""))

        birth_date = _normalize_birth_with_fallback(rec.get("birth_raw", ""), rec.get("birth_date"))
        if isinstance(birth_date, dt.date):
            _safe_set(ws, r, 3, _format_birth_iso(birth_date))
        else:
            _safe_set(ws, r, 3, rec.get("birth_raw", ""))

        address_raw = rec.get("address_raw", rec.get("address", ""))
        address = rec.get("address_api") or _normalize_address(address_raw)
        _safe_set(ws, r, 4, address)

        _safe_set(ws, r, 5, rec.get("father_name", ""))
        _safe_set(ws, r, 6, rec.get("mother_name", ""))
        _safe_set(ws, r, 7, rec.get("father_phone", ""))
        _safe_set(ws, r, 8, rec.get("mother_phone", ""))
        _safe_set(ws, r, 9, _clean_siblings(rec.get("siblings", "")))

        bus = ""
        if rec.get("boarding_method") == "학교차량이용":
            bus = rec.get("boarding_vehicle", "")
        _safe_set(ws, r, 10, bus)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()