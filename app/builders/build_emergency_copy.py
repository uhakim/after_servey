from io import BytesIO
from collections import defaultdict

import openpyxl

from app.mapper import make_student_id


def _vehicle_sort_key(v):
    s = str(v or "")
    digits = "".join(ch for ch in s if ch.isdigit())
    return (int(digits) if digits else 9999, s)


def build_emergency_copy(records):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "전체목록_세로형"
    ws1.append(["차량", "승차위치", "학번", "이름", "전화번호"])

    rows = []
    for r in records:
        if r.get("boarding_method") != "학교차량이용":
            continue
        vehicle = r.get("boarding_vehicle", "")
        location = r.get("boarding_location", "")
        sid = make_student_id(r.get("grade_num"), r.get("class_num"), r.get("number"))
        phone = r.get("mother_phone") or r.get("main_parent_phone") or ""
        rows.append((vehicle, location, sid, r.get("name", ""), phone))

    rows.sort(key=lambda x: (_vehicle_sort_key(x[0]), x[1], x[2]))
    for row in rows:
        ws1.append(list(row))

    for col, width in [(1, 14), (2, 26), (3, 10), (4, 12), (5, 16)]:
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    ws2 = wb.create_sheet("승차위치별_복붙")
    ws2.append(["승차위치", "학번", "이름", "전화번호"])

    grouped = defaultdict(list)
    for vehicle, location, sid, name, phone in rows:
        grouped[location].append((sid, name, phone, vehicle))

    line = 2
    for location in sorted(grouped.keys()):
        ws2.cell(line, 1).value = location
        line += 1
        for sid, name, phone, vehicle in sorted(grouped[location], key=lambda x: x[0]):
            ws2.cell(line, 2).value = sid
            ws2.cell(line, 3).value = name
            ws2.cell(line, 4).value = phone
            ws2.cell(line, 5).value = vehicle
            line += 1
        line += 1

    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 16
    ws2.column_dimensions["E"].width = 12

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
