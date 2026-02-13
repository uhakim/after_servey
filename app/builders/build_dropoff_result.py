from io import BytesIO

import openpyxl
from openpyxl.styles import PatternFill, Border, Side

from app.mapper import WEEKDAYS, make_student_id


DAY_COLS = {
    "월요일": (4, 5, 6, 7),
    "화요일": (8, 9, 10, 11),
    "수요일": (12, 13, 14, 15),
    "목요일": (16, 17, 18, 19),
    "금요일": (20, 21, 22, 23),
}

HEADERS = [
    "학반",
    "학번",
    "이름",
    "월요일 하교방법",
    "월요일 하교시간",
    "탑승차량",
    "하차장소",
    "화요일 하교방법",
    "화요일 하교시간",
    "탑승차량",
    "하차장소",
    "수요일 하교방법",
    "수요일 하교시간",
    "탑승차량",
    "하차장소",
    "목요일 하교방법",
    "목요일 하교시간",
    "탑승차량",
    "하차장소",
    "금요일 하교방법",
    "금요일 하교시간",
    "탑승차량",
    "하차장소",
]


def _to_int(v):
    if v in (None, ""):
        return None
    s = "".join(ch for ch in str(v) if ch.isdigit())
    if not s:
        return None
    return int(s)


def _read_roster_order_from_template(template_bytes=None, default_template_path=None):
    if not template_bytes and not default_template_path:
        return []

    wb = openpyxl.load_workbook(BytesIO(template_bytes)) if template_bytes else openpyxl.load_workbook(default_template_path)
    if "명단" not in wb.sheetnames:
        return []

    ws = wb["명단"]
    rows = []
    for rr in range(3, ws.max_row + 1):
        number = _to_int(ws.cell(rr, 1).value)
        name = str(ws.cell(rr, 2).value or "").strip()
        if number is None and not name:
            continue
        rows.append((number, name))
    return rows


def build_dropoff_result(records, template_bytes=None, default_template_path=None):
    rec_by_number = {r.get("number"): r for r in records if r.get("number") is not None}
    rec_by_name = {r.get("name"): r for r in records}

    grade_num = records[0].get("grade_num") if records else 0
    class_num = records[0].get("class_num") if records else 0
    class_code = f"{grade_num}{class_num}" if grade_num and class_num else ""

    roster_rows = _read_roster_order_from_template(template_bytes, default_template_path)
    if not roster_rows:
        roster_rows = sorted([(r.get("number"), r.get("name", "")) for r in records], key=lambda x: (x[0] or 9999, x[1]))

    wb = openpyxl.Workbook()
    ws_result = wb.active
    ws_result.title = "하교차량조사결과"

    ws_result.append(HEADERS)

    for number, name in roster_rows:
        row = [""] * len(HEADERS)
        row[0] = class_code
        row[1] = make_student_id(grade_num, class_num, number) if number else ""
        row[2] = name

        rec = rec_by_number.get(number) or rec_by_name.get(name)
        if rec:
            for day in WEEKDAYS:
                c_method, c_time, c_vehicle, c_loc = DAY_COLS[day]
                d = rec.get("dropoff", {}).get(day, {})
                row[c_method - 1] = d.get("method", "")
                row[c_time - 1] = d.get("time", "")
                row[c_vehicle - 1] = d.get("vehicle", "")
                row[c_loc - 1] = d.get("location", "")

        ws_result.append(row)

    day_fills = {
        "월요일": PatternFill(fill_type="solid", fgColor="00FFF2CC"),
        "화요일": PatternFill(fill_type="solid", fgColor="00E2F0D9"),
        "수요일": PatternFill(fill_type="solid", fgColor="00D9E1F2"),
        "목요일": PatternFill(fill_type="solid", fgColor="00FCE4D6"),
        "금요일": PatternFill(fill_type="solid", fgColor="00E4DFEC"),
    }
    day_cols = {
        "월요일": range(4, 8),
        "화요일": range(8, 12),
        "수요일": range(12, 16),
        "목요일": range(16, 20),
        "금요일": range(20, 24),
    }
    for day, cols in day_cols.items():
        fill = day_fills[day]
        for r in range(1, ws_result.max_row + 1):
            for c in cols:
                ws_result.cell(r, c).fill = fill

    # Draw full grid borders with normal thin lines.
    thin = Side(style="thin", color="000000")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(1, ws_result.max_row + 1):
        for c in range(1, len(HEADERS) + 1):
            ws_result.cell(r, c).border = grid

    widths = {
        1: 8, 2: 8, 3: 12,
        4: 14, 5: 12, 6: 12, 7: 20,
        8: 14, 9: 12, 10: 12, 11: 20,
        12: 14, 13: 12, 14: 12, 15: 20,
        16: 14, 17: 12, 18: 12, 19: 20,
        20: 14, 21: 12, 22: 12, 23: 20,
    }
    for c, w in widths.items():
        ws_result.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
