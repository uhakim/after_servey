from io import BytesIO
from collections import defaultdict

import openpyxl
from openpyxl.styles import Border, Side

from app.mapper import make_student_id


def _vehicle_sort_key(v):
    s = str(v or "")
    digits = "".join(ch for ch in s if ch.isdigit())
    return (int(digits) if digits else 9999, s)


def build_boarding_report(records):
    wb = openpyxl.Workbook()

    # 상세
    ws_detail = wb.active
    ws_detail.title = "등교차량_상세"
    ws_detail.append(["차량", "승차위치", "학번", "이름", "전화번호"])

    detail_rows = []
    for r in records:
        if r.get("boarding_method") != "학교차량이용":
            continue
        vehicle = r.get("boarding_vehicle", "")
        location = r.get("boarding_location", "")
        sid = make_student_id(r.get("grade_num"), r.get("class_num"), r.get("number"))
        phone = r.get("mother_phone") or r.get("main_parent_phone") or ""
        detail_rows.append((vehicle, location, sid, r.get("name", ""), phone))

    detail_rows.sort(key=lambda x: (_vehicle_sort_key(x[0]), x[1], x[2]))
    for row in detail_rows:
        ws_detail.append(list(row))

    for col, width in [(1, 14), (2, 26), (3, 10), (4, 12), (5, 16)]:
        ws_detail.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # 요약
    ws_summary = wb.create_sheet("등교차량_요약")
    locations = sorted({x[1] for x in detail_rows})
    vehicles = sorted({x[0] for x in detail_rows}, key=_vehicle_sort_key)

    ws_summary.cell(1, 1).value = "승차위치"
    for i, v in enumerate(vehicles, start=2):
        ws_summary.cell(1, i).value = v
    ws_summary.cell(1, len(vehicles) + 2).value = "합계"

    for r_idx, loc in enumerate(locations, start=2):
        ws_summary.cell(r_idx, 1).value = loc
        row_total = 0
        for c_idx, v in enumerate(vehicles, start=2):
            n = sum(1 for row in detail_rows if row[0] == v and row[1] == loc)
            ws_summary.cell(r_idx, c_idx).value = n
            row_total += n
        ws_summary.cell(r_idx, len(vehicles) + 2).value = row_total

    ws_summary.column_dimensions["A"].width = 26
    for c in range(2, len(vehicles) + 3):
        ws_summary.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 12

    # 위치별 명단
    ws_block = wb.create_sheet("등교차량_위치별명단")
    ws_block.append(["차량", "승차위치", "학번", "이름", "전화번호"])

    grouped = defaultdict(list)
    for row in detail_rows:
        grouped[row[0]].append(row)

    line = 2
    thick = Side(style="thick", color="000000")

    for vehicle in sorted(grouped.keys(), key=_vehicle_sort_key):
        vehicle_rows = sorted(grouped[vehicle], key=lambda x: (x[1], x[2]))

        # 호차 시작 구분선
        for c in range(1, 6):
            ws_block.cell(line, c).border = Border(top=thick)

        for v, loc, sid, name, phone in vehicle_rows:
            ws_block.cell(line, 1).value = v
            ws_block.cell(line, 2).value = loc
            ws_block.cell(line, 3).value = sid
            ws_block.cell(line, 4).value = name
            ws_block.cell(line, 5).value = phone
            line += 1

        line += 1

    ws_block.column_dimensions["A"].width = 12
    ws_block.column_dimensions["B"].width = 26
    ws_block.column_dimensions["C"].width = 10
    ws_block.column_dimensions["D"].width = 12
    ws_block.column_dimensions["E"].width = 16

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
